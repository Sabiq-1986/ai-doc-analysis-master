"""
Excel Parser - Comprehensive Excel workbook extraction with full feature support.
Handles .xlsx/.xlsm via openpyxl (dual-load for formulas + values) and .xls via xlrd.
Extracts merged cells, hidden rows/columns, grouping, frozen panes, print areas,
charts, pivot tables, sparklines, images, comments, hyperlinks, data validation,
conditional formatting, named ranges, structured tables, VBA detection,
currency formatting (25+ currencies), formulas with computed values,
header detection heuristics, column statistics, and full workbook metadata.
Returns List[langchain_core.documents.Document] with standardized metadata.
"""
import io
import logging
import math
import os
import re
import zipfile
from datetime import datetime, date, time, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from xml.etree import ElementTree

from langchain_core.documents import Document

from app.config import get_settings

settings = get_settings()
logger = logging.getLogger(__name__)

# Optional dependency: xlrd for legacy .xls files
try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False
    logger.info("[Excel] xlrd not installed; .xls support disabled")

# Currency symbols for 25+ currencies
CURRENCY_SYMBOLS = {
    'AED': ('د.إ', 'AED'), 'SAR': ('﷼', 'ر.س', 'SAR'), 'QAR': ('ر.ق', 'QAR'),
    'BHD': ('BD', 'BHD'), 'KWD': ('د.ك', 'KWD'), 'OMR': ('ر.ع', 'OMR'),
    'EGP': ('ج.م', 'E£', 'EGP'), 'JOD': ('د.ا', 'JOD'),
    'USD': ('$', 'USD'), 'EUR': ('€', 'EUR'), 'GBP': ('£', 'GBP'),
    'INR': ('₹', 'INR'), 'CNY': ('¥', 'CN¥', 'CNY'), 'JPY': ('¥', 'JP¥', 'JPY'),
    'CHF': ('CHF', 'Fr'), 'CAD': ('C$', 'CA$', 'CAD'), 'AUD': ('A$', 'AU$', 'AUD'),
    'SGD': ('S$', 'SGD'), 'HKD': ('HK$', 'HKD'), 'NZD': ('NZ$', 'NZD'),
    'TRY': ('₺', 'TL', 'TRY'), 'MYR': ('RM', 'MYR'), 'THB': ('฿', 'THB'),
    'PHP': ('₱', 'PHP'), 'ZAR': ('R', 'ZAR'),
}

# Flat lookup: symbol -> currency code
_SYMBOL_TO_CURRENCY: Dict[str, str] = {}
for _code, _syms in CURRENCY_SYMBOLS.items():
    for _sym in _syms:
        if _sym not in _SYMBOL_TO_CURRENCY:
            _SYMBOL_TO_CURRENCY[_sym] = _code

EXCEL_ERRORS = {'#REF!', '#N/A', '#VALUE!', '#DIV/0!', '#NAME?', '#NULL!', '#NUM!'}

TOTALS_PATTERNS = re.compile(
    r'\b(total|sum|subtotal|grand\s*total|المجموع|الإجمالي|مجموع)\b', re.IGNORECASE
)


class ExcelParser:
    """
    Production-quality Excel parser supporting .xlsx, .xlsm, and .xls formats.
    Uses openpyxl with dual-load strategy (data_only=True for values, False for formulas).
    Falls back to xlrd for legacy .xls files. Returns LangChain Document objects.
    """

    def __init__(self, image_parser=None):
        """Args: image_parser: Optional ImageParser instance for OCR of embedded images."""
        self.image_parser = image_parser

    # === PUBLIC API ========================================================

    def parse(self, file_path: str) -> List[Document]:
        """Parse an Excel file and return structured LangChain Documents."""
        file_path = str(file_path)
        filename = Path(file_path).name
        ext = Path(file_path).suffix.lower()
        try:
            file_size = os.path.getsize(file_path)
        except OSError:
            file_size = 0
        logger.info(f"[Excel] Parsing {filename} ({file_size} bytes)")
        if ext == '.xls':
            return self._parse_xls(file_path, filename, file_size)
        elif ext in ('.xlsx', '.xlsm'):
            return self._parse_openpyxl(file_path, filename, file_size, ext)
        else:
            logger.warning(f"[Excel] Unsupported extension: {ext}")
            return []

    # === OPENPYXL PARSING (.xlsx / .xlsm) ==================================

    def _parse_openpyxl(self, file_path: str, filename: str,
                        file_size: int, ext: str) -> List[Document]:
        """Parse .xlsx/.xlsm with openpyxl dual-load strategy."""
        import openpyxl
        from openpyxl.utils import get_column_letter

        documents: List[Document] = []
        try:
            wb_values = openpyxl.load_workbook(file_path, data_only=True, read_only=False, keep_links=True)
        except Exception as e:
            logger.error(f"[Excel] Failed to load workbook (values): {e}")
            return []
        try:
            wb_formulas = openpyxl.load_workbook(file_path, data_only=False, read_only=False, keep_links=True)
        except Exception as e:
            logger.warning(f"[Excel] Failed to load workbook (formulas): {e}")
            wb_formulas = None

        # Workbook-level metadata
        props = wb_values.properties
        has_vba = ext == '.xlsm' or (hasattr(wb_values, 'vba_archive') and wb_values.vba_archive is not None)
        workbook_meta = {
            "title": props.title or "", "author": props.creator or "",
            "company": getattr(props, 'company', '') or "",
            "manager": getattr(props, 'manager', '') or "",
            "created": str(props.created) if props.created else "",
            "modified": str(props.modified) if props.modified else "",
            "last_modified_by": props.lastModifiedBy or "",
            "sheet_count": len(wb_values.sheetnames),
            "sheets": wb_values.sheetnames,
            "has_vba_macros": has_vba, "file_size": file_size,
        }
        named_ranges_info = self._extract_named_ranges(wb_values)
        meta_lines = [f"[Workbook: {filename}]"]
        for key, val in workbook_meta.items():
            if val and val not in (0, False, '', []):
                meta_lines.append(f"  {key}: {val}")
        if named_ranges_info:
            meta_lines.append("  Named Ranges:")
            for nr in named_ranges_info:
                meta_lines.append(f"    {nr}")
        if has_vba:
            meta_lines.append("  [VBA MACROS DETECTED]")
        documents.append(Document(
            page_content='\n'.join(meta_lines),
            metadata={"source": filename, "type": "excel", "section": "workbook_properties", **workbook_meta}
        ))

        embedded_images = self._extract_embedded_images(file_path)
        all_formulas: List[str] = []
        all_cross_sheet: List[str] = []

        # Process each sheet
        for sheet_name in wb_values.sheetnames:
            ws_val = wb_values[sheet_name]
            ws_frm = wb_formulas[sheet_name] if wb_formulas else None

            # Sheet visibility
            visibility = "visible"
            try:
                st = ws_val.sheet_state
                if st == 'hidden': visibility = "hidden"
                elif st == 'veryHidden': visibility = "veryHidden"
            except Exception: pass

            # Sheet tab color
            tab_color = None
            try:
                if ws_val.sheet_properties and ws_val.sheet_properties.tabColor:
                    tc = ws_val.sheet_properties.tabColor
                    tab_color = tc.rgb if tc.rgb else str(tc.theme)
            except Exception: pass

            # Frozen panes
            freeze_info = str(ws_val.freeze_panes) if getattr(ws_val, 'freeze_panes', None) else None

            # Print area and print titles
            print_area = str(ws_val.print_area) if getattr(ws_val, 'print_area', None) else None
            print_titles = None
            try:
                parts = []
                if ws_val.print_title_rows: parts.append(f"Rows: {ws_val.print_title_rows}")
                if ws_val.print_title_cols: parts.append(f"Cols: {ws_val.print_title_cols}")
                if parts: print_titles = ', '.join(parts)
            except Exception: pass

            # Page setup (orientation, paper size)
            page_setup_info = None
            try:
                ps = ws_val.page_setup
                if ps:
                    orientation = getattr(ps, 'orientation', None) or 'default'
                    paper = getattr(ps, 'paperSize', None)
                    page_setup_info = f"orientation={orientation}" + (f", paper_size={paper}" if paper else "")
            except Exception: pass

            # Sheet protection status
            protection_status = False
            try:
                if ws_val.protection and ws_val.protection.sheet: protection_status = True
            except Exception: pass

            # Auto-filters
            auto_filter_info = None
            try:
                if ws_val.auto_filter and ws_val.auto_filter.ref:
                    af_ref = ws_val.auto_filter.ref
                    filtered_cols = []
                    if ws_val.auto_filter.filterColumn:
                        for fc in ws_val.auto_filter.filterColumn:
                            filtered_cols.append(str(fc.colId))
                    auto_filter_info = f"Range: {af_ref}"
                    if filtered_cols: auto_filter_info += f", Filtered: {', '.join(filtered_cols)}"
            except Exception: pass

            merged_map = self._build_merged_map(ws_val)
            hidden_rows = self._detect_hidden_rows(ws_val)
            hidden_cols = self._detect_hidden_cols(ws_val)
            row_outlines = self._detect_row_outlines(ws_val)
            col_outlines = self._detect_col_outlines(ws_val)
            comments_list = self._extract_comments(ws_val, sheet_name)
            hyperlinks_list = self._extract_hyperlinks(ws_val, sheet_name)
            validations_list = self._extract_data_validations(ws_val, sheet_name)
            cond_fmt_list = self._extract_conditional_formatting(ws_val, sheet_name)
            charts_list = self._extract_charts(ws_val, sheet_name)
            tables_list = self._extract_tables(ws_val, sheet_name)
            sparklines_list = self._extract_sparklines(ws_val, sheet_name)

            max_row = ws_val.max_row or 0
            max_col = ws_val.max_column or 0
            if max_row == 0:
                continue

            header_rows, headers = self._detect_headers(ws_val, max_col)
            totals_row_idx = self._detect_totals_row(ws_val, max_row, max_col)

            # Sheet summary document
            sl = [f"[Sheet: {sheet_name}]", f"  Visibility: {visibility}",
                  f"  Dimensions: {max_row} rows x {max_col} columns"]
            if tab_color: sl.append(f"  Tab color: {tab_color}")
            if freeze_info: sl.append(f"  Frozen panes: {freeze_info}")
            if print_area: sl.append(f"  Print area: {print_area}")
            if print_titles: sl.append(f"  Print titles: {print_titles}")
            if page_setup_info: sl.append(f"  Page setup: {page_setup_info}")
            if protection_status: sl.append("  [SHEET PROTECTED]")
            if auto_filter_info: sl.append(f"  Auto-filter: {auto_filter_info}")
            if hidden_rows: sl.append(f"  Hidden rows: {len(hidden_rows)}")
            if hidden_cols: sl.append(f"  Hidden columns: {len(hidden_cols)}")
            if row_outlines: sl.append(f"  Row outlines: {len(row_outlines)} groups")
            if col_outlines: sl.append(f"  Column outlines: {len(col_outlines)} groups")
            if headers: sl.append(f"  Headers: {' | '.join(str(h) for h in headers)}")
            if totals_row_idx: sl.append(f"  Totals row: row {totals_row_idx}")
            if comments_list: sl.append(f"  Comments: {len(comments_list)}")
            if hyperlinks_list: sl.append(f"  Hyperlinks: {len(hyperlinks_list)}")
            if validations_list: sl.append(f"  Data validations: {len(validations_list)}")
            if cond_fmt_list: sl.append(f"  Conditional formatting: {len(cond_fmt_list)}")
            if charts_list: sl.append(f"  Charts: {len(charts_list)}")
            if tables_list: sl.append(f"  Tables: {len(tables_list)}")
            if sparklines_list: sl.append(f"  Sparklines: {len(sparklines_list)}")
            documents.append(Document(
                page_content='\n'.join(sl),
                metadata={"source": filename, "type": "excel", "sheet": sheet_name,
                          "section": "sheet_summary", "visibility": visibility}
            ))

            # Row-by-row data extraction
            sheet_col_data: Dict[int, List[float]] = {}
            for row_idx in range(1, max_row + 1):
                if row_idx in header_rows:
                    continue
                row_is_hidden = row_idx in hidden_rows
                row_outline_level = row_outlines.get(row_idx, 0)
                is_totals = (row_idx == totals_row_idx)
                cell_parts: List[str] = []
                row_has_content = False

                for col_idx in range(1, max_col + 1):
                    col_letter = get_column_letter(col_idx)
                    cell_ref = f"{col_letter}{row_idx}"
                    col_is_hidden = col_idx in hidden_cols
                    val_cell = ws_val.cell(row=row_idx, column=col_idx)
                    frm_cell = ws_frm.cell(row=row_idx, column=col_idx) if ws_frm else None

                    merged_info = self._get_merged_info(row_idx, col_idx, merged_map)
                    raw_value = val_cell.value
                    formula_expr = None
                    if frm_cell and isinstance(frm_cell.value, str) and frm_cell.value.startswith('='):
                        formula_expr = frm_cell.value

                    if raw_value is None and formula_expr is None and merged_info is None:
                        continue
                    if merged_info is not None:
                        if merged_info.get('is_secondary'):
                            continue
                        raw_value = merged_info.get('value', raw_value)

                    formatted = self._format_cell_value(val_cell, raw_value, formula_expr)
                    if not formatted:
                        continue
                    row_has_content = True

                    header_label = ''
                    if headers and col_idx <= len(headers) and headers[col_idx - 1]:
                        header_label = f"{headers[col_idx - 1]}: "

                    cell_text = f"{header_label}{formatted}"
                    annotations = []
                    if merged_info and merged_info.get('span'): annotations.append(merged_info['span'])
                    if col_is_hidden: annotations.append("[HIDDEN COL]")
                    if row_is_hidden: annotations.append("[HIDDEN ROW]")
                    if row_outline_level > 0: annotations.append(f"[outline L{row_outline_level}]")
                    if is_totals: annotations.append("[TOTALS]")
                    try:
                        if val_cell.font and val_cell.font.bold: annotations.append("[bold]")
                    except Exception: pass
                    if annotations:
                        cell_text += ' ' + ' '.join(annotations)
                    cell_parts.append(cell_text)

                    # Collect formula info
                    if formula_expr:
                        fl = f"{sheet_name}!{cell_ref}: {self._safe_str(raw_value)} <- {formula_expr}"
                        if formula_expr.startswith('{=') and formula_expr.endswith('}'):
                            fl += " [ARRAY]"
                        all_formulas.append(fl)
                        if self._has_cross_sheet_ref(formula_expr, sheet_name):
                            all_cross_sheet.append(fl)

                    # Collect numeric data for column statistics
                    if isinstance(raw_value, (int, float)) and not isinstance(raw_value, bool):
                        if not math.isnan(raw_value) and not math.isinf(raw_value):
                            sheet_col_data.setdefault(col_idx, []).append(raw_value)

                if row_has_content and cell_parts:
                    row_text = ' | '.join(cell_parts)

                    # Compute row-level numeric aggregations
                    row_numbers = []
                    for col_idx in range(1, max_col + 1):
                        val = ws_val.cell(row=row_idx, column=col_idx).value
                        if isinstance(val, (int, float)) and not isinstance(val, bool):
                            if not math.isnan(val) and not math.isinf(val):
                                row_numbers.append(val)

                    # Add row summary with totals if there are numeric values
                    if row_numbers and len(row_numbers) >= 2:
                        row_sum = sum(row_numbers)
                        row_count = len(row_numbers)
                        row_avg = row_sum / row_count
                        row_text += f" | [ROW TOTALS: sum={row_sum:.0f}, count={row_count}, avg={row_avg:.1f}]"

                    row_meta = {"source": filename, "type": "excel", "sheet": sheet_name, "row": row_idx}
                    if row_is_hidden:
                        row_text = f"[HIDDEN] {row_text}"
                        row_meta["hidden"] = True
                    if is_totals:
                        row_meta["is_totals_row"] = True
                    documents.append(Document(page_content=row_text, metadata=row_meta))

            # Annotation documents for this sheet
            if comments_list:
                documents.append(Document(
                    page_content=f"[Comments - {sheet_name}]\n" + '\n'.join(comments_list),
                    metadata={"source": filename, "type": "excel", "sheet": sheet_name, "section": "comments"}
                ))
            if hyperlinks_list:
                documents.append(Document(
                    page_content=f"[Hyperlinks - {sheet_name}]\n" + '\n'.join(hyperlinks_list),
                    metadata={"source": filename, "type": "excel", "sheet": sheet_name, "section": "hyperlinks"}
                ))
            if validations_list:
                documents.append(Document(
                    page_content=f"[Data Validations - {sheet_name}]\n" + '\n'.join(validations_list),
                    metadata={"source": filename, "type": "excel", "sheet": sheet_name, "section": "data_validations"}
                ))
            if cond_fmt_list:
                documents.append(Document(
                    page_content=f"[Conditional Formatting - {sheet_name}]\n" + '\n'.join(cond_fmt_list),
                    metadata={"source": filename, "type": "excel", "sheet": sheet_name, "section": "conditional_formatting"}
                ))
            for chart_info in charts_list:
                documents.append(Document(
                    page_content=chart_info,
                    metadata={"source": filename, "type": "excel", "sheet": sheet_name, "section": "chart"}
                ))
            for table_info in tables_list:
                documents.append(Document(
                    page_content=table_info,
                    metadata={"source": filename, "type": "excel", "sheet": sheet_name, "section": "table"}
                ))
            if sparklines_list:
                documents.append(Document(
                    page_content=f"[Sparklines - {sheet_name}]\n" + '\n'.join(sparklines_list),
                    metadata={"source": filename, "type": "excel", "sheet": sheet_name, "section": "sparklines"}
                ))

            # Column statistics
            if sheet_col_data:
                stats_lines = [f"[Column Statistics - {sheet_name}]"]
                for ci in sorted(sheet_col_data.keys()):
                    vals = sheet_col_data[ci]
                    if len(vals) < 2: continue
                    cl = get_column_letter(ci)
                    ch = f" ({headers[ci-1]})" if headers and ci <= len(headers) and headers[ci-1] else ""
                    stats_lines.append(
                        f"  Column {cl}{ch}: count={len(vals)}, sum={sum(vals):.2f}, "
                        f"avg={sum(vals)/len(vals):.2f}, min={min(vals):.2f}, max={max(vals):.2f}"
                    )
                if len(stats_lines) > 1:
                    documents.append(Document(
                        page_content='\n'.join(stats_lines),
                        metadata={"source": filename, "type": "excel", "sheet": sheet_name, "section": "statistics"}
                    ))

        # Pivot tables from XML
        documents.extend(self._extract_pivot_tables_from_xml(file_path, filename))

        # Embedded images OCR
        if embedded_images and self.image_parser:
            for img_info in embedded_images:
                try:
                    img_data = img_info.get('data', b'')
                    img_name = img_info.get('name', 'embedded_image')
                    if not img_data: continue
                    ocr_text = self.image_parser.ocr_image_bytes(img_data, img_name)
                    if ocr_text and ocr_text.strip():
                        documents.append(Document(
                            page_content=f"[Embedded Image: {img_name}]\n{ocr_text}",
                            metadata={"source": filename, "type": "excel",
                                      "sheet": img_info.get('sheet', 'unknown'),
                                      "section": "embedded_image", "image_name": img_name}
                        ))
                except Exception as e:
                    logger.debug(f"[Excel] Image OCR failed: {e}")

        # Formulas summary
        if all_formulas:
            fl = [f"[Formulas Summary - {filename}]", f"  Total formulas: {len(all_formulas)}"]
            if all_cross_sheet:
                fl.append(f"  Cross-sheet references: {len(all_cross_sheet)}")
                fl.append("  Cross-sheet formulas:")
                for csf in all_cross_sheet[:50]: fl.append(f"    {csf}")
            fl.append("  All formulas:")
            for f_line in all_formulas[:200]: fl.append(f"    {f_line}")
            documents.append(Document(
                page_content='\n'.join(fl),
                metadata={"source": filename, "type": "excel", "section": "formulas_summary",
                          "formula_count": len(all_formulas), "cross_sheet_count": len(all_cross_sheet)}
            ))

        try: wb_values.close()
        except Exception: pass
        if wb_formulas:
            try: wb_formulas.close()
            except Exception: pass

        logger.info(f"[Excel] Extracted {len(documents)} documents from {filename}")
        return documents

    # === MERGED CELLS ======================================================

    def _build_merged_map(self, ws) -> Dict[Tuple[int, int], Dict]:
        """Build mapping from every cell in a merged range to its merge info."""
        merged_map: Dict[Tuple[int, int], Dict] = {}
        try:
            for mr in ws.merged_cells.ranges:
                row_span = mr.max_row - mr.min_row + 1
                col_span = mr.max_col - mr.min_col + 1
                span_label = f"[merged {row_span}x{col_span}]"
                try: origin_value = ws.cell(row=mr.min_row, column=mr.min_col).value
                except Exception: origin_value = None
                for r in range(mr.min_row, mr.max_row + 1):
                    for c in range(mr.min_col, mr.max_col + 1):
                        is_origin = (r == mr.min_row and c == mr.min_col)
                        merged_map[(r, c)] = {
                            'is_origin': is_origin, 'is_secondary': not is_origin,
                            'value': origin_value, 'span': span_label,
                        }
        except Exception as e:
            logger.debug(f"[Excel] Merged cells error: {e}")
        return merged_map

    @staticmethod
    def _get_merged_info(row: int, col: int, merged_map: Dict) -> Optional[Dict]:
        """Return merged cell info if the cell is part of a merged range."""
        return merged_map.get((row, col))

    # === HIDDEN ROWS & COLUMNS =============================================

    def _detect_hidden_rows(self, ws) -> set:
        hidden = set()
        try:
            for row_idx, rd in ws.row_dimensions.items():
                if rd.hidden: hidden.add(row_idx)
        except Exception as e:
            logger.debug(f"[Excel] Hidden rows error: {e}")
        return hidden

    def _detect_hidden_cols(self, ws) -> set:
        hidden = set()
        try:
            from openpyxl.utils import column_index_from_string
            for col_key, cd in ws.column_dimensions.items():
                if cd.hidden:
                    hidden.add(column_index_from_string(col_key) if isinstance(col_key, str) else int(col_key))
        except Exception as e:
            logger.debug(f"[Excel] Hidden cols error: {e}")
        return hidden

    # === ROW & COLUMN GROUPING / OUTLINES ==================================

    def _detect_row_outlines(self, ws) -> Dict[int, int]:
        outlines: Dict[int, int] = {}
        try:
            for row_idx, rd in ws.row_dimensions.items():
                level = getattr(rd, 'outlineLevel', 0) or 0
                if level > 0: outlines[row_idx] = level
        except Exception as e:
            logger.debug(f"[Excel] Row outlines error: {e}")
        return outlines

    def _detect_col_outlines(self, ws) -> Dict[int, int]:
        outlines: Dict[int, int] = {}
        try:
            from openpyxl.utils import column_index_from_string
            for col_key, cd in ws.column_dimensions.items():
                level = getattr(cd, 'outlineLevel', 0) or 0
                if level > 0:
                    idx = column_index_from_string(col_key) if isinstance(col_key, str) else int(col_key)
                    outlines[idx] = level
        except Exception as e:
            logger.debug(f"[Excel] Column outlines error: {e}")
        return outlines

    # === COMMENTS / NOTES ==================================================

    def _extract_comments(self, ws, sheet_name: str) -> List[str]:
        results = []
        try:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.comment:
                        author = cell.comment.author or "Unknown"
                        text = (cell.comment.text or "").strip()
                        if text: results.append(f"  {cell.coordinate}: [{author}] {text}")
        except Exception as e:
            logger.debug(f"[Excel] Comments error on {sheet_name}: {e}")
        return results

    # === HYPERLINKS ========================================================

    def _extract_hyperlinks(self, ws, sheet_name: str) -> List[str]:
        results = []
        try:
            if hasattr(ws, 'hyperlinks') and ws.hyperlinks:
                for hl in ws.hyperlinks:
                    ref, target = hl.ref or "?", hl.target or ""
                    display, tooltip = hl.display or "", getattr(hl, 'tooltip', '') or ""
                    line = f"  {ref}: {target}"
                    if display: line += f" (display: {display})"
                    if tooltip: line += f" (tooltip: {tooltip})"
                    results.append(line)
        except Exception as e:
            logger.debug(f"[Excel] Hyperlinks error on {sheet_name}: {e}")
        try:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.hyperlink and cell.hyperlink.target:
                        line = f"  {cell.coordinate}: {cell.hyperlink.target}"
                        disp = self._safe_str(cell.value)
                        if disp: line += f" (text: {disp})"
                        if line not in results: results.append(line)
        except Exception as e:
            logger.debug(f"[Excel] Cell hyperlinks error on {sheet_name}: {e}")
        return results

    # === DATA VALIDATIONS ==================================================

    def _extract_data_validations(self, ws, sheet_name: str) -> List[str]:
        results = []
        try:
            if not hasattr(ws, 'data_validations') or not ws.data_validations:
                return results
            dvs = ws.data_validations.dataValidation
            if not dvs: return results
            for dv in dvs:
                cells = str(dv.sqref) if dv.sqref else "?"
                line = f"  Cells: {cells}, Type: {dv.type or 'none'}"
                if dv.operator: line += f", Op: {dv.operator}"
                if dv.formula1: line += f", Source: {dv.formula1}"
                if dv.formula2: line += f", Formula2: {dv.formula2}"
                if dv.allow_blank is not None: line += f", AllowBlank: {dv.allow_blank}"
                show_dd = getattr(dv, 'showDropDown', None)
                if show_dd is not None: line += f", Dropdown: {not show_dd}"
                pt, pm = getattr(dv, 'promptTitle', '') or "", getattr(dv, 'prompt', '') or ""
                if pt or pm: line += f", Prompt: [{pt}] {pm}"
                et = getattr(dv, 'errorTitle', '') or ""
                em, es = getattr(dv, 'error', '') or "", getattr(dv, 'errorStyle', '') or ""
                if et or em: line += f", Error: [{es}] [{et}] {em}"
                results.append(line)
        except Exception as e:
            logger.debug(f"[Excel] Data validation error on {sheet_name}: {e}")
        return results

    # === CONDITIONAL FORMATTING ============================================

    def _extract_conditional_formatting(self, ws, sheet_name: str) -> List[str]:
        results = []
        try:
            if not ws.conditional_formatting: return results
            for cf_rule_list in ws.conditional_formatting:
                cells_range = str(cf_rule_list)
                try: rules = cf_rule_list.rules
                except AttributeError: rules = []
                for rule in rules:
                    rt = getattr(rule, 'type', 'unknown')
                    pri = getattr(rule, 'priority', '')
                    op = getattr(rule, 'operator', '') or ''
                    formulas = []
                    if hasattr(rule, 'formula') and rule.formula:
                        formulas = [str(f) for f in rule.formula]
                    line = f"  Range: {cells_range}, Type: {rt}"
                    if pri: line += f", Priority: {pri}"
                    if op: line += f", Op: {op}"
                    if formulas: line += f", Formula: {'; '.join(formulas)}"
                    # Color scale
                    if hasattr(rule, 'colorScale') and rule.colorScale:
                        cs = rule.colorScale
                        colors = [getattr(c, 'rgb', None) or str(getattr(c, 'theme', ''))
                                  for c in (cs.color or [])]
                        line += f", ColorScale: {colors}"
                    # Data bars
                    if hasattr(rule, 'dataBar') and rule.dataBar:
                        db = rule.dataBar
                        fc = getattr(db.color, 'rgb', str(db.color)) if hasattr(db, 'color') and db.color else ""
                        line += f", DataBar: color={fc}"
                    # Icon sets
                    if hasattr(rule, 'iconSet') and rule.iconSet:
                        line += f", IconSet: {getattr(rule.iconSet, 'iconSet', 'unknown')}"
                    # DXF details
                    if hasattr(rule, 'dxf') and rule.dxf:
                        dxf = rule.dxf
                        dp = []
                        if dxf.font:
                            fd = []
                            if dxf.font.bold: fd.append("bold")
                            if dxf.font.italic: fd.append("italic")
                            if dxf.font.color and dxf.font.color.rgb: fd.append(f"color={dxf.font.color.rgb}")
                            if fd: dp.append(f"font=[{','.join(fd)}]")
                        if dxf.fill:
                            fc = ""
                            if dxf.fill.fgColor and dxf.fill.fgColor.rgb: fc = dxf.fill.fgColor.rgb
                            elif dxf.fill.bgColor and dxf.fill.bgColor.rgb: fc = dxf.fill.bgColor.rgb
                            if fc: dp.append(f"fill={fc}")
                        if dxf.border: dp.append("border=yes")
                        if dp: line += f", DXF: {'; '.join(dp)}"
                    results.append(line)
        except Exception as e:
            logger.debug(f"[Excel] Conditional formatting error on {sheet_name}: {e}")
        return results

    # === CHARTS ============================================================

    def _extract_charts(self, ws, sheet_name: str) -> List[str]:
        results = []
        try:
            if not hasattr(ws, '_charts') or not ws._charts: return results
            for idx, chart in enumerate(ws._charts):
                lines = [f"[Chart {idx+1} - {sheet_name}]", f"  Type: {type(chart).__name__}"]
                try:
                    if chart.title:
                        lines.append(f"  Title: {chart.title if isinstance(chart.title, str) else str(chart.title)}")
                except Exception: pass
                try:
                    if hasattr(chart, 'style') and chart.style: lines.append(f"  Style: {chart.style}")
                except Exception: pass
                try:
                    if hasattr(chart, 'x_axis') and chart.x_axis:
                        xt = getattr(chart.x_axis, 'title', None)
                        if xt: lines.append(f"  X-Axis: {xt}")
                    if hasattr(chart, 'y_axis') and chart.y_axis:
                        yt = getattr(chart.y_axis, 'title', None)
                        if yt: lines.append(f"  Y-Axis: {yt}")
                except Exception: pass
                try:
                    if hasattr(chart, 'series') and chart.series:
                        for si, series in enumerate(chart.series):
                            sl = f"  Series {si+1}:"
                            st = getattr(series, 'title', None) or getattr(series, 'tx', None)
                            if st: sl += f" {st}"
                            vr = getattr(series, 'val', None)
                            if vr and hasattr(vr, 'numRef') and vr.numRef:
                                sl += f" Values={vr.numRef.f}"
                                if vr.numRef.numCache and vr.numRef.numCache.pt:
                                    pts = [str(pt.v) for pt in vr.numRef.numCache.pt[:10]]
                                    sl += f" Data=[{', '.join(pts)}]"
                                    if len(vr.numRef.numCache.pt) > 10: sl += "..."
                            cr = getattr(series, 'cat', None)
                            if cr:
                                if hasattr(cr, 'numRef') and cr.numRef: sl += f" Categories={cr.numRef.f}"
                                elif hasattr(cr, 'strRef') and cr.strRef: sl += f" Categories={cr.strRef.f}"
                            lines.append(sl)
                except Exception as e:
                    logger.debug(f"[Excel] Chart series error: {e}")
                results.append('\n'.join(lines))
        except Exception as e:
            logger.debug(f"[Excel] Charts error on {sheet_name}: {e}")
        return results

    # === STRUCTURED TABLES / LIST OBJECTS ==================================

    def _extract_tables(self, ws, sheet_name: str) -> List[str]:
        results = []
        try:
            if not hasattr(ws, 'tables') or not ws.tables: return results
            for tname, table in ws.tables.items():
                lines = [f"[Table: {table.displayName or tname}]", f"  Range: {table.ref}"]
                if hasattr(table, 'tableStyleInfo') and table.tableStyleInfo:
                    sn = getattr(table.tableStyleInfo, 'name', '')
                    if sn: lines.append(f"  Style: {sn}")
                if hasattr(table, 'tableColumns') and table.tableColumns:
                    col_names, calc_cols = [], []
                    for tc in table.tableColumns.tableColumn:
                        col_names.append(tc.name)
                        if hasattr(tc, 'calculatedColumnFormula') and tc.calculatedColumnFormula:
                            ft = getattr(tc.calculatedColumnFormula, 'text', '') or \
                                 getattr(tc.calculatedColumnFormula, 'attr_text', '') or ""
                            if ft: calc_cols.append(f"{tc.name}: {ft}")
                    lines.append(f"  Columns: {', '.join(col_names)}")
                    if calc_cols:
                        lines.append("  Calculated Columns:")
                        for cc in calc_cols: lines.append(f"    {cc}")
                if hasattr(table, 'totalsRowShown') and table.totalsRowShown:
                    lines.append("  Totals Row: Yes")
                    if hasattr(table, 'tableColumns') and table.tableColumns:
                        for tc in table.tableColumns.tableColumn:
                            tf = getattr(tc, 'totalsRowFunction', None)
                            if tf: lines.append(f"    {tc.name}: {tf}")
                results.append('\n'.join(lines))
        except Exception as e:
            logger.debug(f"[Excel] Tables error on {sheet_name}: {e}")
        return results

    # === SPARKLINES ========================================================

    def _extract_sparklines(self, ws, sheet_name: str) -> List[str]:
        results = []
        try:
            sg = getattr(ws, 'sparkline_groups', None) or getattr(ws, '_sparkline_groups', None)
            if not sg: return results
            for group in sg:
                sp_type = getattr(group, 'type', 'line') or 'line'
                for sp in (getattr(group, 'sparklines', []) or []):
                    sq = getattr(sp, 'sqref', '') or ''
                    dr = getattr(sp, 'f', '') or ''
                    results.append(f"  Type: {sp_type}, Cell: {sq}, Data: {dr}")
        except Exception as e:
            logger.debug(f"[Excel] Sparklines error on {sheet_name}: {e}")
        return results

    # === PIVOT TABLES (FROM XML) ===========================================

    def _extract_pivot_tables_from_xml(self, file_path: str, filename: str) -> List[Document]:
        documents = []
        try:
            if not zipfile.is_zipfile(file_path): return documents
            with zipfile.ZipFile(file_path, 'r') as zf:
                pivot_files = [n for n in zf.namelist() if 'pivotTable' in n and n.endswith('.xml')]
                for pf in pivot_files:
                    try:
                        root = ElementTree.fromstring(zf.read(pf))
                        ns = (root.tag.split('}')[0] + '}') if root.tag.startswith('{') else ''
                        pivot_name = root.attrib.get('name', Path(pf).stem)
                        lines = [f"[Pivot Table: {pivot_name}]"]
                        loc = root.find(f'{ns}location')
                        if loc is not None:
                            ref = loc.attrib.get('ref', '')
                            if ref: lines.append(f"  Location: {ref}")
                            fdr = loc.attrib.get('firstDataRow', '')
                            fdc = loc.attrib.get('firstDataCol', '')
                            if fdr or fdc: lines.append(f"  First data: row={fdr}, col={fdc}")
                        for tag, label in [('rowFields', 'Row fields'), ('colFields', 'Column fields')]:
                            el = root.find(f'{ns}{tag}')
                            if el is not None:
                                idxs = [f.attrib.get('x', '?') for f in el.findall(f'{ns}field')]
                                lines.append(f"  {label}: {', '.join(idxs)}")
                        df_el = root.find(f'{ns}dataFields')
                        if df_el is not None:
                            for df in df_el.findall(f'{ns}dataField'):
                                lines.append(
                                    f"  Data field: {df.attrib.get('name', '?')} "
                                    f"(function={df.attrib.get('subtotal', 'sum')}, "
                                    f"field_index={df.attrib.get('fld', '?')})"
                                )
                        pf_el = root.find(f'{ns}pageFields')
                        if pf_el is not None:
                            for p in pf_el.findall(f'{ns}pageField'):
                                lines.append(f"  Filter: {p.attrib.get('name', '')} (field={p.attrib.get('fld', '?')})")
                        documents.append(Document(
                            page_content='\n'.join(lines),
                            metadata={"source": filename, "type": "excel", "section": "pivot_table", "pivot_name": pivot_name}
                        ))
                    except Exception as e:
                        logger.debug(f"[Excel] Pivot XML parse error ({pf}): {e}")
                # Pivot cache definitions
                for cf_path in [n for n in zf.namelist() if 'pivotCacheDefinition' in n and n.endswith('.xml')]:
                    try:
                        root = ElementTree.fromstring(zf.read(cf_path))
                        ns = (root.tag.split('}')[0] + '}') if root.tag.startswith('{') else ''
                        cfs = root.find(f'{ns}cacheFields')
                        if cfs is not None:
                            fnames = []
                            for cf in cfs.findall(f'{ns}cacheField'):
                                n = cf.attrib.get('name', '?')
                                nf = cf.attrib.get('numFmtId', '')
                                fnames.append(n + (f" (fmt={nf})" if nf else ""))
                            if fnames:
                                documents.append(Document(
                                    page_content=f"[Pivot Cache: {Path(cf_path).stem}]\n  Fields: {', '.join(fnames)}",
                                    metadata={"source": filename, "type": "excel", "section": "pivot_cache"}
                                ))
                    except Exception as e:
                        logger.debug(f"[Excel] Pivot cache error ({cf_path}): {e}")
        except Exception as e:
            logger.debug(f"[Excel] Pivot extraction error: {e}")
        return documents

    # === EMBEDDED IMAGES ===================================================

    def _extract_embedded_images(self, file_path: str) -> List[Dict]:
        images = []
        try:
            if not zipfile.is_zipfile(file_path): return images
            img_exts = {'.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tiff', '.emf', '.wmf'}
            with zipfile.ZipFile(file_path, 'r') as zf:
                for entry in zf.namelist():
                    if entry.startswith('xl/media/') and Path(entry).suffix.lower() in img_exts:
                        try:
                            data = zf.read(entry)
                            images.append({'name': Path(entry).name, 'data': data, 'path': entry, 'sheet': 'unknown', 'size': len(data)})
                        except Exception as e:
                            logger.debug(f"[Excel] Image read error ({entry}): {e}")
                # Map images to sheets via drawing relationships
                sheet_map = {}
                for dr in [n for n in zf.namelist() if 'drawings' in n and n.endswith('.rels')]:
                    try:
                        root = ElementTree.fromstring(zf.read(dr))
                        for rel in root:
                            target = rel.attrib.get('Target', '')
                            if '../media/' in target:
                                sheet_map[Path(target).name] = Path(dr).parent.name
                    except Exception: pass
                for img in images:
                    if img['name'] in sheet_map: img['sheet'] = sheet_map[img['name']]
        except Exception as e:
            logger.debug(f"[Excel] Image extraction error: {e}")
        return images

    # === NAMED RANGES ======================================================

    def _extract_named_ranges(self, wb) -> List[str]:
        results = []
        try:
            if hasattr(wb, 'defined_names'):
                for dn in wb.defined_names.definedName:
                    name = dn.name
                    value = dn.attr_text if hasattr(dn, 'attr_text') else str(dn.value)
                    scope = "Global"
                    if dn.localSheetId is not None:
                        try: scope = f"Sheet: {wb.sheetnames[dn.localSheetId]}"
                        except (IndexError, TypeError): scope = f"Sheet ID: {dn.localSheetId}"
                    line = f"{name} = {value} [{scope}]"
                    if getattr(dn, 'hidden', False): line += " [HIDDEN]"
                    results.append(line)
        except Exception as e:
            logger.debug(f"[Excel] Named ranges error: {e}")
        return results

    # === HEADER DETECTION HEURISTICS =======================================

    def _detect_headers(self, ws, max_col: int) -> Tuple[set, List[str]]:
        """Detect header row(s) using bold, fill color, and string ratio heuristics."""
        header_rows: set = set()
        headers: List[str] = []
        max_check = min(10, ws.max_row or 1)
        if max_check == 0: return header_rows, headers

        row_scores: Dict[int, float] = {}
        for row_idx in range(1, max_check + 1):
            bold_count = fill_count = string_count = number_count = non_empty = 0
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is None: continue
                non_empty += 1
                try:
                    if cell.font and cell.font.bold: bold_count += 1
                except Exception: pass
                try:
                    if cell.fill and cell.fill.fgColor:
                        fg = cell.fill.fgColor
                        if (fg.rgb and fg.rgb not in ('00000000', 'FFFFFFFF')) or fg.theme is not None:
                            fill_count += 1
                except Exception: pass
                if isinstance(cell.value, str): string_count += 1
                elif isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool): number_count += 1
            if non_empty == 0: continue
            score = 0.0
            if bold_count / non_empty >= 0.5: score += 3.0
            if fill_count / non_empty >= 0.3: score += 2.0
            sr = string_count / non_empty
            if sr >= 0.7: score += 2.0
            elif sr >= 0.5: score += 1.0
            if row_idx == 1: score += 1.0
            row_scores[row_idx] = score

        if not row_scores: return header_rows, headers
        best = max(row_scores.values())
        threshold = max(3.0, best * 0.6)
        for ri, sc in row_scores.items():
            if sc >= threshold: header_rows.add(ri)
        if not header_rows and 1 in row_scores and row_scores[1] >= 1.5:
            header_rows.add(1)

        if header_rows:
            sorted_hr = sorted(header_rows)
            for col_idx in range(1, max_col + 1):
                parts = []
                for hr in sorted_hr:
                    v = ws.cell(row=hr, column=col_idx).value
                    if v is not None: parts.append(self._safe_str(v).strip())
                headers.append(' / '.join(p for p in parts if p) or "")
        return header_rows, headers

    # === TOTALS ROW DETECTION ==============================================

    def _detect_totals_row(self, ws, max_row: int, max_col: int) -> Optional[int]:
        if max_row < 3: return None
        for row_idx in range(max_row, max(0, max_row - 5), -1):
            for col_idx in range(1, min(max_col + 1, 6)):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val and TOTALS_PATTERNS.search(self._safe_str(val)):
                    return row_idx
        return None

    # === CELL VALUE FORMATTING =============================================

    def _format_cell_value(self, cell, raw_value, formula_expr: Optional[str]) -> Optional[str]:
        """Format cell with type awareness, currency, formulas, errors, dates."""
        if raw_value is None and formula_expr is None:
            return None

        def _append_formula(result: str) -> str:
            if not formula_expr: return result
            if formula_expr.startswith('{=') and formula_expr.endswith('}'):
                return f"{result} <- {formula_expr} [ARRAY]"
            return f"{result} <- {formula_expr}"

        if isinstance(raw_value, str) and raw_value in EXCEL_ERRORS:
            return _append_formula(f"[ERROR: {raw_value}]")
        if isinstance(raw_value, bool):
            return _append_formula("TRUE" if raw_value else "FALSE")
        if isinstance(raw_value, datetime):
            return _append_formula(raw_value.strftime('%Y-%m-%d %H:%M:%S'))
        if isinstance(raw_value, date):
            return _append_formula(raw_value.strftime('%Y-%m-%d'))
        if isinstance(raw_value, time):
            return _append_formula(raw_value.strftime('%H:%M:%S'))
        if isinstance(raw_value, timedelta):
            ts = int(raw_value.total_seconds())
            h, rem = divmod(abs(ts), 3600)
            m, s = divmod(rem, 60)
            return _append_formula(f"{'-' if ts < 0 else ''}{h:02d}:{m:02d}:{s:02d}")
        if isinstance(raw_value, (int, float)):
            nf = None
            try: nf = cell.number_format
            except Exception: pass
            return _append_formula(self._format_number(raw_value, nf))
        if isinstance(raw_value, str):
            r = raw_value.strip()
            if formula_expr and formula_expr != raw_value:
                return _append_formula(r)
            return r if r else None
        r = self._safe_str(raw_value)
        return _append_formula(r) if r else None

    def _format_number(self, value: float, num_format: Optional[str]) -> str:
        """Format numeric value: currencies, percentages, scientific, accounting, fractions, dates."""
        if num_format is None or num_format == 'General':
            return str(int(value)) if isinstance(value, int) or (isinstance(value, float) and value == int(value)) else f"{value:g}"

        fmt_lower = num_format.lower()

        # Percentage
        if '%' in num_format:
            dm = re.search(r'0\.(0+)', num_format)
            d = len(dm.group(1)) if dm else 1
            return f"{value * 100:.{d}f}%"

        # Scientific notation
        if 'e+' in fmt_lower or 'e-' in fmt_lower:
            dm = re.search(r'0\.(0+)', num_format)
            return f"{value:.{len(dm.group(1)) if dm else 2}E}"

        # Fraction
        if '/' in num_format and '#' in num_format:
            return self._format_fraction(value)

        # Date (numbers stored as dates)
        date_ind = ['yyyy', 'yy', 'mm', 'dd', 'mmm', 'mmmm', 'h', 'hh', 'ss', 'am/pm', 'a/p']
        if any(d in fmt_lower for d in date_ind):
            return self._serial_to_date(value, num_format)

        # Currency from format string
        cc = self._detect_currency_from_format(num_format)
        if cc:
            sym = CURRENCY_SYMBOLS.get(cc, (cc,))[0]
            dm = re.search(r'0\.(0+)', num_format)
            d = len(dm.group(1)) if dm else 2
            if '_(' in num_format or '(' in num_format:
                return f"({sym}{abs(value):,.{d}f})" if value < 0 else f"{sym}{value:,.{d}f}"
            return f"{sym}{value:,.{d}f} {cc}"

        # Accounting without currency
        if '_(' in num_format or '_(* ' in num_format:
            dm = re.search(r'0\.(0+)', num_format)
            d = len(dm.group(1)) if dm else 2
            return f"({abs(value):,.{d}f})" if value < 0 else f"{value:,.{d}f}"

        # Comma/thousands
        if '#,##' in num_format or '#,#' in num_format:
            dm = re.search(r'0\.(0+)', num_format)
            d = len(dm.group(1)) if dm else 0
            return f"{value:,.{d}f}"

        return str(int(value)) if isinstance(value, int) or (isinstance(value, float) and value == int(value)) else f"{value:g}"

    def _detect_currency_from_format(self, num_format: str) -> Optional[str]:
        """Detect currency from Excel number format string ([$symbol-locale] or direct)."""
        if not num_format: return None
        lm = re.search(r'\[\$(.+?)\-[0-9A-Fa-f]+\]', num_format)
        if lm:
            sym = lm.group(1).strip()
            if sym in _SYMBOL_TO_CURRENCY: return _SYMBOL_TO_CURRENCY[sym]
            for s, c in _SYMBOL_TO_CURRENCY.items():
                if s in sym or sym in s: return c
        for sym in sorted(_SYMBOL_TO_CURRENCY.keys(), key=len, reverse=True):
            if sym in num_format: return _SYMBOL_TO_CURRENCY[sym]
        return None

    def _format_fraction(self, value: float) -> str:
        if value == 0: return "0"
        sign = '-' if value < 0 else ''
        av = abs(value)
        whole = int(av)
        frac = av - whole
        if frac < 1e-9: return f"{sign}{whole}"
        best_n, best_d, min_err = 0, 1, frac
        for den in range(2, 101):
            num = round(frac * den)
            err = abs(frac - num / den)
            if err < min_err:
                min_err, best_n, best_d = err, num, den
                if err < 1e-9: break
        if best_n == 0: return f"{sign}{whole}" if whole else "0"
        return f"{sign}{whole} {best_n}/{best_d}" if whole > 0 else f"{sign}{best_n}/{best_d}"

    def _serial_to_date(self, value: float, num_format: str) -> str:
        try:
            serial = int(value)
            frac = value - serial
            if serial < 1:
                ts = int(round(frac * 86400))
                h, rem = divmod(ts, 3600)
                m, s = divmod(rem, 60)
                return f"{h:02d}:{m:02d}:{s:02d}"
            if serial >= 60: serial -= 1
            dt = datetime(1899, 12, 31) + timedelta(days=serial)
            if frac > 0: dt += timedelta(seconds=int(round(frac * 86400)))
            fl = num_format.lower()
            has_d = any(d in fl for d in ['yyyy', 'yy', 'mm', 'dd', 'mmm', 'mmmm'])
            has_t = any(t in fl for t in ['h', 'hh', 'ss', 'am/pm'])
            if has_d and has_t: return dt.strftime('%Y-%m-%d %H:%M:%S')
            elif has_t and not has_d: return dt.strftime('%H:%M:%S')
            return dt.strftime('%Y-%m-%d')
        except Exception:
            return str(value)

    # === CROSS-SHEET FORMULA DETECTION =====================================

    @staticmethod
    def _has_cross_sheet_ref(formula: str, current_sheet: str) -> bool:
        if not formula: return False
        for m in re.findall(r"(?:'[^']+'|[A-Za-z0-9_]+)!", formula):
            ref = m.rstrip('!').strip("'")
            if ref.lower() != current_sheet.lower(): return True
        return False

    # === XLS PARSING (LEGACY FORMAT VIA xlrd) ==============================

    def _parse_xls(self, file_path: str, filename: str, file_size: int) -> List[Document]:
        if not XLRD_AVAILABLE:
            logger.warning(f"[Excel] xlrd not installed, cannot parse .xls: {filename}")
            return [Document(
                page_content=f"[Excel File: {filename}]\nCannot parse .xls: xlrd not installed.",
                metadata={"source": filename, "type": "excel", "error": "xlrd_not_available"}
            )]
        documents: List[Document] = []
        try:
            wb = xlrd.open_workbook(file_path, formatting_info=True)
        except Exception:
            try: wb = xlrd.open_workbook(file_path, formatting_info=False)
            except Exception as e2:
                logger.error(f"[Excel] xlrd failed: {e2}")
                return [Document(page_content=f"[Excel: {filename}]\nFailed: {e2}",
                                 metadata={"source": filename, "type": "excel", "error": str(e2)})]

        # Workbook metadata
        meta_lines = [f"[Workbook: {filename}]", "  Format: XLS (Legacy)",
                      f"  Sheets: {wb.nsheets}", f"  Sheet names: {', '.join(wb.sheet_names())}",
                      f"  File size: {file_size}"]
        documents.append(Document(
            page_content='\n'.join(meta_lines),
            metadata={"source": filename, "type": "excel", "section": "workbook_properties",
                      "sheet_count": wb.nsheets, "file_size": file_size}
        ))

        # Named ranges
        try:
            if wb.name_obj_list:
                nr_lines = ["[Named Ranges]"]
                for nr in wb.name_obj_list:
                    scope = "Global"
                    if nr.scope >= 0:
                        try: scope = f"Sheet: {wb.sheet_names()[nr.scope]}"
                        except IndexError: scope = f"Sheet ID: {nr.scope}"
                    nr_lines.append(f"  {nr.name} [{scope}]")
                if len(nr_lines) > 1:
                    documents.append(Document(
                        page_content='\n'.join(nr_lines),
                        metadata={"source": filename, "type": "excel", "section": "named_ranges"}
                    ))
        except Exception as e:
            logger.debug(f"[Excel] xlrd named ranges error: {e}")

        # Process each sheet
        for sheet_idx in range(wb.nsheets):
            ws = wb.sheet_by_index(sheet_idx)
            sheet_name = ws.name
            visibility = "visible"
            try:
                vv = wb.sheet_visibility(sheet_idx)
                if vv == 1: visibility = "hidden"
                elif vv == 2: visibility = "veryHidden"
            except Exception: pass

            summary = [f"[Sheet: {sheet_name}]", f"  Visibility: {visibility}",
                       f"  Dimensions: {ws.nrows} rows x {ws.ncols} columns"]

            # Merged cells tracking
            xls_merged = set()
            try:
                for rlo, rhi, clo, chi in ws.merged_cells:
                    for r in range(rlo, rhi):
                        for c in range(clo, chi):
                            if not (r == rlo and c == clo): xls_merged.add((r, c))
            except Exception: pass

            documents.append(Document(
                page_content='\n'.join(summary),
                metadata={"source": filename, "type": "excel", "sheet": sheet_name,
                          "section": "sheet_summary", "visibility": visibility}
            ))

            xls_headers = self._detect_headers_xls(ws)
            for row_idx in range(ws.nrows):
                if row_idx == 0 and xls_headers: continue
                cell_parts = []
                has_content = False
                for col_idx in range(ws.ncols):
                    if (row_idx, col_idx) in xls_merged: continue
                    ct = ws.cell_type(row_idx, col_idx)
                    cv = ws.cell_value(row_idx, col_idx)
                    if ct == xlrd.XL_CELL_EMPTY: continue
                    has_content = True
                    fmt = self._format_xls_cell(ct, cv, wb.datemode)
                    if not fmt: continue
                    hl = f"{xls_headers[col_idx]}: " if xls_headers and col_idx < len(xls_headers) and xls_headers[col_idx] else ""
                    cell_parts.append(f"{hl}{fmt}")
                if has_content and cell_parts:
                    rt = ' | '.join(cell_parts)
                    rm = {"source": filename, "type": "excel", "sheet": sheet_name, "row": row_idx + 1}
                    if visibility != "visible":
                        rt = f"[{visibility.upper()} SHEET] {rt}"
                        rm["visibility"] = visibility
                    documents.append(Document(page_content=rt, metadata=rm))

            # Hyperlinks
            try:
                if hasattr(ws, 'hyperlink_map') and ws.hyperlink_map:
                    hl_lines = [f"[Hyperlinks - {sheet_name}]"]
                    for (r, c), hl in ws.hyperlink_map.items():
                        target = getattr(hl, 'url_or_path', '') or ''
                        desc = getattr(hl, 'desc', '') or ''
                        line = f"  Row {r+1}, Col {c+1}: {target}"
                        if desc: line += f" (text: {desc})"
                        hl_lines.append(line)
                    if len(hl_lines) > 1:
                        documents.append(Document(
                            page_content='\n'.join(hl_lines),
                            metadata={"source": filename, "type": "excel", "sheet": sheet_name, "section": "hyperlinks"}
                        ))
            except Exception as e:
                logger.debug(f"[Excel] xlrd hyperlinks error on {sheet_name}: {e}")

        wb.release_resources()
        logger.info(f"[Excel] Extracted {len(documents)} documents from {filename} (xls)")
        return documents

    # === XLS HEADER DETECTION ==============================================

    @staticmethod
    def _detect_headers_xls(ws) -> List[str]:
        if ws.nrows == 0 or ws.ncols == 0: return []
        headers, string_count, non_empty = [], 0, 0
        for col_idx in range(ws.ncols):
            ct = ws.cell_type(0, col_idx)
            cv = ws.cell_value(0, col_idx)
            if ct == xlrd.XL_CELL_EMPTY:
                headers.append('')
                continue
            non_empty += 1
            if ct == xlrd.XL_CELL_TEXT: string_count += 1
            headers.append(str(cv).strip())
        return headers if non_empty > 0 and string_count / non_empty >= 0.5 else []

    # === XLS CELL FORMATTING ===============================================

    @staticmethod
    def _format_xls_cell(cell_type: int, cell_value, datemode: int) -> Optional[str]:
        if cell_type == xlrd.XL_CELL_EMPTY: return None
        if cell_type == xlrd.XL_CELL_TEXT:
            t = str(cell_value).strip()
            return f"[ERROR: {t}]" if t in EXCEL_ERRORS else (t or None)
        if cell_type == xlrd.XL_CELL_NUMBER:
            if isinstance(cell_value, float) and cell_value == int(cell_value):
                return str(int(cell_value))
            return f"{cell_value:g}"
        if cell_type == xlrd.XL_CELL_DATE:
            try:
                dt = xlrd.xldate_as_tuple(cell_value, datemode)
                if dt[0] == 0 and dt[1] == 0 and dt[2] == 0:
                    return f"{dt[3]:02d}:{dt[4]:02d}:{dt[5]:02d}"
                if dt[3] == 0 and dt[4] == 0 and dt[5] == 0:
                    return f"{dt[0]:04d}-{dt[1]:02d}-{dt[2]:02d}"
                return f"{dt[0]:04d}-{dt[1]:02d}-{dt[2]:02d} {dt[3]:02d}:{dt[4]:02d}:{dt[5]:02d}"
            except Exception: return str(cell_value)
        if cell_type == xlrd.XL_CELL_BOOLEAN:
            return "TRUE" if cell_value else "FALSE"
        if cell_type == xlrd.XL_CELL_ERROR:
            codes = {0x00: '#NULL!', 0x07: '#DIV/0!', 0x0F: '#VALUE!',
                     0x17: '#REF!', 0x1D: '#NAME?', 0x24: '#NUM!', 0x2A: '#N/A'}
            return f"[ERROR: {codes.get(cell_value, f'#ERROR({cell_value})')}]"
        return str(cell_value) if cell_value else None

    # === UTILITY ===========================================================

    @staticmethod
    def _safe_str(value) -> str:
        if value is None: return ''
        try: return str(value)
        except Exception: return repr(value)
